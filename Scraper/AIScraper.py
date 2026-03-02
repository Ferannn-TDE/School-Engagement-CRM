import asyncio
from abc import ABC, abstractmethod
import re
import threading
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
from SchoolNamePuller import SupabaseSchoolPuller, CsvSchoolPuller

SCRAPE_BATCH_SIZE = 1

class AIModel(ABC):
    @abstractmethod
    async def search(self, facility_name, city, website):
        pass

class GeminiModel(AIModel):
    async def search(self, facility_name, city, website):
        async with async_playwright() as plawright:
            browser = await plawright.chromium.launch(headless=False)
            context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
            page = await context.new_page()
            await page.goto("https://gemini.google.com/app")
            await page.wait_for_timeout(4000)
            search_query = (
                f"Find deans and career counselors contact information from {facility_name} "
                f"in {city}, Illinois. Website: {website}. "
                "The school services at least grades 9-12. "
                "Return results in a table format with columns: Name, Title, Email, Phone, School Name"
            )
            await page.keyboard.type(search_query)
            await page.keyboard.press("Enter")
            await page.wait_for_timeout(4000)
            try:
                await page.wait_for_selector("table", timeout=30000)
            except Exception:
                await browser.close()
                return []
            rows = await page.query_selector_all("table tr")
            contacts = []
            for row_index, row in enumerate(rows):
                if row_index == 0:
                    continue
                cells = await row.query_selector_all("td")
                email = await self._extract_email(cells[2])
                data = {
                    "Name": (await cells[0].text_content() or "").strip(),
                    "Title": (await cells[1].text_content() or "").strip(),
                    "Email": email,
                    "Phone": (await cells[3].text_content() or "").strip(),
                    "School Name": (await cells[4].text_content() or "").strip(),
                }
                contacts.append(data)
            await browser.close()
            await asyncio.sleep(5)
            return contacts
    
    async def _extract_email(self, cell):
        text = (await cell.text_content() or "").strip()
        match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
        if match:
            return match.group(0)        
        try:
            anchors = await cell.query_selector_all("a")
            if anchors:
                href = await anchors[0].get_attribute("href")
                if href and href.lower().startswith("mailto:"):
                    email = href[7:].split("?")[0].strip()
                    if email:
                        return email
        except Exception:
            pass
        return ""


class ChatGPTModel(AIModel):
    async def search(self, facility_name, city, website):
        async with async_playwright() as plawright:
            browser = await plawright.chromium.launch(headless=False)
            context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
            page = await context.new_page()
            await page.goto("https://chatgpt.com")
            await page.wait_for_timeout(4000)
            search_query = (
                f"Find deans and career counselors contact information from {facility_name} "
                f"in {city}, Illinois. Website: {website}. "
                "The school services at least grades 9-12. "
                "Return results in a table format with columns: Name, Title, Email, Phone, School Name"
            )
            await page.keyboard.type(search_query)
            await page.keyboard.press("Enter")
            await page.wait_for_timeout(4000)
            try:
                await page.wait_for_selector("table", timeout=30000)
            except Exception:
                await browser.close()
                return []
            rows = await page.query_selector_all("table tr")
            contacts = []
            for row_index, row in enumerate(rows):
                if row_index == 0:
                    continue
                cells = await row.query_selector_all("td")
                email = await self._extract_email(cells[2])
                data = {
                    "Name": (await cells[0].text_content() or "").strip(),
                    "Title": (await cells[1].text_content() or "").strip(),
                    "Email": email,
                    "Phone": (await cells[3].text_content() or "").strip(),
                    "School Name": (await cells[4].text_content() or "").strip(),
                }
                contacts.append(data)
            await browser.close()
            await asyncio.sleep(5)
            return contacts
    
    async def _extract_email(self, cell):
        text = (await cell.text_content() or "").strip()
        match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
        if match:
            return match.group(0)        
        try:
            anchors = await cell.query_selector_all("a")
            if anchors:
                href = await anchors[0].get_attribute("href")
                if href and href.lower().startswith("mailto:"):
                    email = href[7:].split("?")[0].strip()
                    if email:
                        return email
        except Exception:
            pass
        return ""
    

class CopilotModel(AIModel):
    async def search(self, facility_name, city, website):
        async with async_playwright() as plawright:
            browser = await plawright.chromium.launch(headless=False)
            context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
            page = await context.new_page()
            await page.goto("https://copilot.microsoft.com/")
            await page.wait_for_timeout(4000)
            search_query = (
                f"Find deans and career counselors contact information from {facility_name} "
                f"in {city}, Illinois. Website: {website}. "
                "The school services at least grades 9-12. "
                "Return results in a table format with columns: Name, Title, Email, Phone, School Name"
            )
            await page.keyboard.type(search_query)
            await page.keyboard.press("Enter")
            await page.wait_for_timeout(4000)
            try:
                await page.wait_for_selector("table", timeout=30000)
            except Exception:
                await browser.close()
                return []
            rows = await page.query_selector_all("table tr")
            contacts = []
            for row_index, row in enumerate(rows):
                if row_index == 0:
                    continue
                cells = await row.query_selector_all("td")
                email = await self._extract_email(cells[2])
                data = {
                    "Name": (await cells[0].text_content() or "").strip(),
                    "Title": (await cells[1].text_content() or "").strip(),
                    "Email": email,
                    "Phone": (await cells[3].text_content() or "").strip(),
                    "School Name": (await cells[4].text_content() or "").strip(),
                }
                contacts.append(data)
            await browser.close()
            await asyncio.sleep(5)
            return contacts
    
    async def _extract_email(self, cell):
        text = (await cell.text_content() or "").strip()
        match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text)
        if match:
            return match.group(0)        
        try:
            anchors = await cell.query_selector_all("a")
            if anchors:
                href = await anchors[0].get_attribute("href")
                if href and href.lower().startswith("mailto:"):
                    email = href[7:].split("?")[0].strip()
                    if email:
                        return email
        except Exception:
            pass
        return ""


    
class DocumentWriter(ABC):
    @abstractmethod
    def write(self, records, filename):
        pass


class ExcelWriter(DocumentWriter):
    def __init__(self):
        self.lock = threading.Lock()

    def write(self, contacts, filename):
        if not contacts:
            return
        with self.lock:
            try:
                workbook = load_workbook(filename)
                worksheet = workbook.active
            except FileNotFoundError:
                workbook = Workbook()
                worksheet = workbook.active
            headers = ["Name", "Title", "Email", "Phone", "School Name"]
            first_cell = worksheet.cell(1, 1).value
            if first_cell is None:
                for col, header in enumerate(headers, 1):
                    worksheet.cell(1, col, header)
                start_row = 2
            else:
                start_row = worksheet.max_row + 1
            for row_idx, record in enumerate(contacts, start_row):
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row_idx, col_idx, record.get(header, ""))
            workbook.save(filename)



class Scraper:
    def __init__(self, ai_model, writer, school_puller):
        self.ai_model = ai_model
        self.writer = writer
        self.school_puller = school_puller

    async def scrape_facility(self, facility):
        return await self.ai_model.search(
            facility.get("FacilityName", ""),
            facility.get("City", ""),
            facility.get("Website", ""),
        )

    async def run(self, output_file):
        facilities = self.school_puller.get_schools()
        total_contacts = 0
        for i in range(0, len(facilities), SCRAPE_BATCH_SIZE):
            group = facilities[i:i + SCRAPE_BATCH_SIZE]
            processes = [self.scrape_facility(f) for f in group]
            scrape_data = await asyncio.gather(*processes, return_exceptions=True)
            contacts = []
            for result in scrape_data:
                if isinstance(result, Exception):
                    continue
                contacts.extend(result)
            self.writer.write(contacts, output_file)
            total_contacts += len(contacts)
            print(f"Processing Schools {i + 1}-{i + len(group)}: {len(contacts)} written")
        print(f"Total contacts: {total_contacts}")

async def main():
    ai_model = GeminiModel()
    writer = ExcelWriter()
    school_puller = SupabaseSchoolPuller()
    scraper = Scraper(ai_model, writer, school_puller)
    await scraper.run("contacts.xlsx")


if __name__ == "__main__":
    asyncio.run(main())